import json
import os
from datetime import datetime, timedelta
from openpyxl import load_workbook, Workbook
from models.modules import normal_function as nf

class ExcelExporter:
    def __init__(self):        
        self.template_file = nf.get_export_template_path()        
        self.report_folder = nf.get_output_folder()
        self.server_data_folder = nf.get_server_folder()

        
    @staticmethod
    def load_json_data(filepath):
        with open(filepath, 'r') as file:
            return json.load(file)

    @staticmethod
    def get_week_dates(year, week_number):
        # Calculate the first day of the year
        first_day_of_year = datetime(year, 1, 1)
        # Adjust for the first Monday of the year
        first_monday = first_day_of_year + timedelta(days=(first_day_of_year.weekday()))
        # Calculate the start of the given week number
        start_date = first_monday + timedelta(weeks=week_number - 1)

        # Generate dates for the week
        week_dates = [start_date + timedelta(days=i) for i in range(5)]  # Monday to Friday
        return [date.strftime('%d/%m/%Y') for date in week_dates]

    @staticmethod
    def merge_json_data(json_data):
        tasks_data = {}
        # Process each week's data
        for week_year_str, week_data in json_data['weeks'].items():
            year = int(week_year_str[2:])  # Get the last four digits as year
            week = int(week_year_str[:2])  # Get the first two digits as week number
            week_dates = ExcelExporter.get_week_dates(year, week)

            # Correct the debug print to handle strings directly
            print(f"Debug: Processing week {week} of {year}, dates are {week_dates}")

            # Process each task category in the week
            for task_category, tasks in week_data.items():
                if task_category == "Week_status":
                    continue  # Skip processing week status
                
                # Initialize task category if not exists
                if task_category not in tasks_data:
                    tasks_data[task_category] = {}

                for task in tasks:
                    # Create a unique identifier for each task based on its attributes
                    task_key = tuple(task[item] for item in ['project', 'category'])
                    if task_key not in tasks_data[task_category]:
                        tasks_data[task_category][task_key] = {**task, 'task_key': task_key, 'totalEffort': 0}
                        for date in week_dates:
                            date_key = datetime.strptime(date, '%d/%m/%Y').strftime('%m/%d')  # Adjust for proper formatting
                            tasks_data[task_category][task_key][date_key] = 0  # Initialize date keys with 0 effort

                    current_task = tasks_data[task_category][task_key]
                    # Update status to the most recent one
                    current_task['status'] = task['status']
                    for i, day in enumerate(['Mon', 'Tue', 'Wed', 'Thu', 'Fri']):
                        date_key = datetime.strptime(week_dates[i], '%d/%m/%Y').strftime('%m/%d')
                        if date_key not in current_task:
                            current_task[date_key] = 0  # Initialize if not already present
                        current_task[date_key] += float(task.get(day, 0))
                        # print(f"Debug: {task_category}, {task_key}, {date_key} -> {current_task[date_key]}")

                    # Update non-daily values
                    current_task.update({k: task[k] for k in task if k not in ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'status', 'totalEffort']})
                    current_task['totalEffort'] += float(task.get('totalEffort', 0))

        # Flatten the tasks for each category into a list, grouped by unique task_keys
        final_output = {}
        for category, tasks_by_key in tasks_data.items():
            final_output[category] = [task for task in tasks_by_key.values()]

        return final_output

    # def export_all_jsons_to_excel(self, json_folder):
    #         # Load the template workbook, which includes the 'temp' sheet
    #         book = load_workbook(self.template_file)
            
    #         # If the workbook already has sheets other than 'temp', remove them
    #         for sheet in book.sheetnames:
    #             if sheet != 'temp':
    #                 std_sheet = book[sheet]
    #                 book.remove(std_sheet)
            
    #         # Define a list of files to exclude
    #         exclude_files = ['Common Account.json', 'config_user.py']

    #         # Iterate through each JSON file in the folder
    #         for filename in os.listdir(json_folder):
    #             if filename.endswith('.json') and filename not in exclude_files:
    #                 member_name = filename.split('.')[0]
    #                 json_path = os.path.join(json_folder, filename)
    #                 json_data = self.load_json_data(json_path)
    #                 merged_tasks = self.merge_json_data(json_data)
    #                 # Copy the 'temp' sheet and rename it
    #                 member_sheet = book.copy_worksheet(book['temp'])
    #                 member_sheet.title = member_name  # Rename the copied sheet to the member's name
                    
    #                 print(f"Debug: {member_name}")
                    
    #                 # Export data to the newly created sheet
    #                 self.json_to_sheet(merged_tasks, member_sheet, member_name)

    #         # Once all sheets are processed, remove the original 'temp' sheet if not needed
    #         if 'temp' in book.sheetnames:
    #             book.remove(book['temp'])

    #         # Save the workbook
    #         book.save(os.path.join(self.report_path, self.output_file_name))
    #         book.close()

    def export_all_jsons_to_excel(self):
        # Load the template workbook
        book = load_workbook(self.template_file)

        # List of template sheets
        template_sheets = [book[sheet_name] for sheet_name in book.sheetnames if 'temp' in sheet_name]

        # Keep track of used sheets
        used_sheets = []

        # Iterate through each JSON file and corresponding template sheet
        for idx, filename in enumerate(os.listdir(self.server_data_folder)):
            if filename.endswith('.json') and filename not in ['Common Account.json', 'config_user.py']:
                if idx >= len(template_sheets):
                    break  # Stop if there are more files than sheets
                member_name = filename.split('.')[0]
                json_path = os.path.join(self.server_data_folder, filename)
                json_data = self.load_json_data(json_path)
                merged_tasks = self.merge_json_data(json_data)

                # Use the corresponding template sheet
                member_sheet = template_sheets[idx]
                member_sheet.title = member_name  # Rename the sheet
                used_sheets.append(member_sheet.title)  # Mark this sheet as used

                print(f"Debug: Processing {member_name}")
                # Export data to the renamed sheet
                self.json_to_sheet(merged_tasks, member_sheet, member_name)

        # Remove any template sheets that were not used
        for sheet in template_sheets:
            if sheet.title not in used_sheets:
                book.remove(sheet)

        # Save the workbook at the specified report path
        # Save the workbook
        day = str(datetime.now().day)
        month = str(datetime.now().month)
        output_file_name = "Report_data" + f"_{day}_{month}.xlsx"
        book.save(os.path.join(self.report_folder, output_file_name))
        book.close()


    def json_to_sheet(self, merged_tasks, sheet, member_name,start_row=9):
        # Write the member name to cell B3
        sheet['B3'] = member_name

        # Assuming 'sheet' is passed as an argument and is already created outside this function
        # Find date columns in the provided 'sheet'
        date_columns = {}
        for cell in sheet[6]:  # Assuming row 6 contains the date headers
            try:
                if isinstance(cell.value, datetime):
                    date_obj = cell.value
                else:
                    date_obj = datetime.strptime(cell.value, '%d-%m-%Y')
                formatted_date = date_obj.strftime('%m/%d')
                date_columns[formatted_date] = cell.column
            except (TypeError, ValueError):
                continue  # Skip cells with non-datetime values or incorrect formatting

        # Convert merged tasks to a list of dictionaries for DataFrame creation
        data_rows = []
        task_number = 1
        for category, tasks in merged_tasks.items():
            for task in tasks:
                row_data = {
                    'No.': task_number,
                    'Project': task['project'],
                    'Task Name': category,
                    'Category': task['category'],
                    'Comment': task['comment'],
                    'Findings': task['findings'],
                    'Requirements': task['requirements'],
                    'Estimated Effort': task['estimatedEffort'],
                    'Start Date': task['startDate'],
                    'End Date': task['endDate'],
                    'Status': task['status']
                }
                # Include the time spent on specific dates from the task
                for date_key in task:
                    if '/' in date_key and date_key in date_columns:
                        row_data[date_columns[date_key]] = task[date_key]

                data_rows.append(row_data)
                task_number += 1

        # Write data to the provided 'sheet'
        for index, row in enumerate(data_rows):
            row_num = start_row + index
            for key, value in row.items():
                if isinstance(key, int):  # If key is an integer, it is a column index
                    sheet.cell(row=row_num, column=key, value=value)
                else:
                    col_index = {'No.': 1, 'Project': 2, 'Task Name': 3, 'Category': 4, 'Comment': 5, 'Findings': 6,
                                'Requirements': 7, 'Estimated Effort': 8, 'Start Date': 10, 'End Date': 11, 'Status': 12}.get(key)
                    if col_index:
                        sheet.cell(row=row_num, column=col_index, value=value)

# Usage
# output_filename = 'Report_Effort_2.xlsx'
# report_folder = r'C:\Users\TAU9HC\Documents\02_Code_practicing\05_TET_tool\tracking_effort_tool\mvc-pattern\output'
# json_folder_path = r'\\bosch.com\dfsRB\DfsVN\LOC\Hc\RBVH\20_EBS\10_EBS3\01_Internal\EBS32_ITK\03_Projects\Automation Tool\Tracking effort Tool\User Data'
# template_file = r'C:\Users\TAU9HC\Documents\02_Code_practicing\05_TET_tool\tracking_effort_tool\mvc-pattern\resources\template\export_template.xlsx'
# exporter = ExcelExporter()
# exporter.export_all_jsons_to_excel()